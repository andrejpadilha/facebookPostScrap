import facebook
import requests
import pprint
import openpyxl
import os
import time

graph = facebook.GraphAPI(access_token='EAACEdEose0cBADO4MV4GEksreoXqFbcrVZCLZCcFXoCwalK9gmz6JCn0SNGx33R9OWzIrtLyxfblHud3W9NdGWZCQXoZCibPfNUMZBHOojQvZBTqHEV1wvuKXUVCf1EFQZBl9AMVan43M1IaayrvBA0jm9k1cTVu9yroCgxWCPq9x8y0XxznZBtZB1gKfvXo1K9wZD', version='2.9')

pp = pprint.PrettyPrinter(indent=4)

os.chdir('..') #muda o diretorio
os.chdir('Desktop') #aqui tambem

wb = openpyxl.load_workbook('infos.xlsx') #abre o arquivo excel com as datas unicode
sheet = wb.get_sheet_by_name('Sheet1') #abre a planilha especifica

wb_output = openpyxl.Workbook()
sheet_output = wb_output.get_active_sheet()
sheet.title = 'Dados Brutos'
sheet_output['A1'] = 'Link'
sheet_output['B1'] = 'Data'
sheet_output['C1'] = 'Compartilhamentos'
sheet_output['D1'] = 'Likes'
sheet_output['E1'] = 'Reacoes (incluindo likes)'
sheet_output['F1'] = 'Comentarios'

range_string = 'F4:F825'
list = []
for row in sheet.iter_rows(range_string):
	for cell in row:
		list.append(cell.value)


for i in range (0,821):
	post = graph.get_object(id='ufabc/feed', fields='limit=100,created_time,shares,likes.limit(0).summary(true),comments.summary(true),reactions.limit(0).summary(true)',since=list[i+1],until=list[i])
	#post = graph.get_object(id='237319529713795')#, fields='limit=100,created_time,shares,likes.limit(0).summary(true),comments.summary(true),reactions.limit(0).summary(true)',since=list[i+1],until=list[i])


	for i in post['data'][:]:
		if 'shares' in i:
			pp.pprint("facebook.com/%s | Criado em: %s. | Compartilhamentos: %s | Likes: %s | Reacoes: %s | Comentarios: %s" % (i['id'],i['created_time'],i['shares']['count'],i['likes']['summary']['total_count'],i['reactions']['summary']['total_count'],i['comments']['summary']['total_count']))
			sheet_output.append(['https://www.facebook.com/%s;%s;%s;%s;%s;%s' % (i['id'],i['created_time'],i['shares']['count'],i['likes']['summary']['total_count'],i['reactions']['summary']['total_count'],i['comments']['summary']['total_count'])])
		else:
			pp.pprint("facebook.com/%s | Criado em: %s. | Compartilhamentos: 0 | Likes: %s | Reacoes: %s | Comentarios: %s" % (i['id'],i['created_time'],i['likes']['summary']['total_count'],i['reactions']['summary']['total_count'],i['comments']['summary']['total_count']))
			sheet_output.append(['https://www.facebook.com/%s;%s;0;%s;%s;%s' % (i['id'],i['created_time'],i['likes']['summary']['total_count'],i['reactions']['summary']['total_count'],i['comments']['summary']['total_count'])])	



#wb_output.save('analiseFacebookUFABC.xlsx')
wb_output.save('analiseFacebookUFABC%s.xlsx' % (time.strftime("%Y,%B,%d %H:%M:%S")))

